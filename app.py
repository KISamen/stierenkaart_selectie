import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# Titel van de app
st.title("🐂 Stieren Data Selectie")

# Upload een Excel-bestand
uploaded_file = st.file_uploader("📂 Upload je Excel-bestand", type=["xlsx"])

# Controleer of er een bestand is geüpload
if uploaded_file is not None:
    try:
        # Laad het Excel-bestand in een Pandas DataFrame, met rij 2 als kolomnamen
        df = pd.read_excel(uploaded_file, engine="openpyxl", header=1)

        # Laat een voorbeeld van de data zien
        st.write("📊 **Voorbeeld van de data:**")
        st.dataframe(df.head())

        # **Stieren staan altijd in kolom C (index 2, want Python begint bij 0)**
        stieren_kolom = df.columns[2]  # Kolom C vastzetten
        ki_kolom = df.columns[1]  # Kolom B bevat de KI-code

        # Haal unieke stierennamen op
        stieren_namen = df[stieren_kolom].dropna().unique()

        # Multiselectie van stierennamen
        geselecteerde_stieren = st.multiselect("🐂 Selecteer de stieren", stieren_namen)

        if geselecteerde_stieren:
            # Filter de data op de geselecteerde stierennamen
            gefilterde_data = df[df[stieren_kolom].isin(geselecteerde_stieren)]

            # **Canada-template volgorde instellen**
            canada_volgorde = {
                "Kicode": "KI Code",
                "Stiernaam": "Bull name",
                "Vader": "Father",
                "Moeders Vader": "Maternal Grandfather",
                "aAa": "aAa",
                "% Betr": "% reliability",
                "Kg melk": "kg milk",
                "% vet": "% fat",
                "% eiwit": "% protein",
                "Kg vet": "kg fat",
                "Kg eiwit": "kg protein",
                "Dcht totaal": "#Daughters",
                "% Betr.1": "% reliability",
                "Frame": "frame",
                "Uier": "udder",
                "Beenwerk": "feet & legs",
                "Totaal exterieur": "final score",
                "Hoogtemaat": "stature",
                "Voorhand": "chest width",
                "Inhoud": "body depth",
                "Openheid": "angularity",
                "Conditie score": "condition score",
                "Kruisligging": "rump angle",
                "Kruisbreedte": "rump width",
                "Beenstand achter": "rear legs rear view",
                "Beenstand zij": "rear leg set",
                "Klauwhoek": "foot angle",
                "Voorbeenstand": "front feet orientation",
                "Beengebruik": "mobility",
                "Vooruieraanhechting": "fore udder attachment",
                "Voorspeenplaatsing": "front teat placement",
                "Speenlengte": "teat length",
                "Uierdiepte": "udder depth",
                "Achteruierhoogte": "rear udder height",
                "Ophangband": "central ligament",
                "Achterspeenplaatsing": "rear teat placement",
                "Uierbalans": "udder balance",
                "Geboortegemak": "calving ease",
                "Melksnelheid": "milking speed",
                "Celgetal": "somatic cell score",
                "Vruchtbaarheid": "female fertility",
                "Karakter": "temperament",
                "Verwantschapsgraad": "maturity rate",
                "Persistentie": "persistence",
                "Klauwgezondheid": "hoof health"
            }

            # Alleen kolommen gebruiken die in de dataset staan
            geselecteerde_kolommen = [col for col in canada_volgorde.keys() if col in df.columns]

            # Categorieën toewijzen aan kolommen
            categorie_mapping = {
                "Production Traits": ["kg milk", "% fat", "% protein", "kg fat", "kg protein", "#Daughters"],
                "Conformation Traits": ["frame", "udder", "feet & legs", "final score"],
                "Lineair Traits": ["stature", "chest width", "body depth", "angularity", "condition score",
                                   "rump angle", "rump width", "rear legs rear view", "rear leg set", "foot angle",
                                   "front feet orientation", "mobility", "fore udder attachment",
                                   "front teat placement", "teat length", "udder depth", "rear udder height",
                                   "central ligament", "rear teat placement", "udder balance"],
                "Management Traits": ["calving ease", "milking speed", "somatic cell score", "female fertility",
                                      "temperament", "maturity rate", "persistence", "hoof health"]
            }

            # Extra regel met categorieën voorbereiden
            categorie_rij = []
            for col in geselecteerde_kolommen:
                categorie = ""
                for cat, cols in categorie_mapping.items():
                    if canada_volgorde[col] in cols:
                        categorie = cat
                        break
                categorie_rij.append(categorie)

            # Dataframe maken
            gefilterde_data = gefilterde_data[geselecteerde_kolommen]
            gefilterde_data = gefilterde_data.rename(columns=canada_volgorde)

            # **Output als Excel (.xlsx) met samengevoegde cellen**
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                gefilterde_data.to_excel(writer, index=False, sheet_name="Data", startrow=1)

                workbook = writer.book
                sheet = writer.sheets["Data"]

                # Schrijf de categorieënrij boven de kolomnamen
                for i, categorie in enumerate(categorie_rij):
                    cell = sheet.cell(row=1, column=i + 1, value=categorie)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Samengevoegde cellen per categorie
                col_indexes = {cat: [] for cat in categorie_mapping.keys()}
                for i, categorie in enumerate(categorie_rij):
                    if categorie:
                        col_indexes[categorie].append(i + 1)

                # Merge de cellen
                for cat, cols in col_indexes.items():
                    if cols:
                        start_col = min(cols)
                        end_col = max(cols)
                        sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

            output.seek(0)

            # **Download-knop voor de Excel-output**
            st.download_button(
                label="⬇️ Download Excel",
                data=output,
                file_name="gesorteerde_data.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            # Laat de gesorteerde data zien
            st.write("✅ **Gesorteerde Data met Categorieën:**")
            st.dataframe(gefilterde_data)

    except Exception as e:
        st.error(f"❌ Er is een fout opgetreden bij het verwerken van het bestand: {e}")

else:
    st.warning("⚠️ Upload een Excel-bestand om te beginnen.")
